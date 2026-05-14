import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from docx import Document
from collections import defaultdict

# ===== 读取源数据 =====

# 1. 读取3月全院质控汇总 - Sheet1 (详细数据)
wb_src1 = openpyxl.load_workbook('3月全院质控汇总.xlsx')
ws_detail = wb_src1['Sheet1']

dept_grade_b = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})
dept_excellent = defaultdict(int)

for row in ws_detail.iter_rows(min_row=2, values_only=True):
    dept = row[3]
    record_type = row[2]
    grade = row[11]

    if not dept:
        continue

    if row[13] == '优秀':
        dept_excellent[dept] += 1

    if grade == '乙级':
        if '门诊' in str(record_type):
            dept_grade_b[dept]['门诊'] += 1
        elif '环节' in str(record_type):
            dept_grade_b[dept]['环节'] += 1
        elif '终末' in str(record_type):
            dept_grade_b[dept]['终末'] += 1

# 2. 读取迟交汇总
wb_late = openpyxl.load_workbook('2026年3月出院病历迟交汇总表.xlsx')
ws_late = wb_late['Sheet1']
dept_late_fine = defaultdict(float)

for row in ws_late.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    dept = row[1]
    fine = row[6] or 0
    dept_late_fine[dept] += fine

# 3. 读取质控医生奖励和不良事件奖励
doc = Document('2026年03月医疗质量安全管理通报.docx')

dept_qc_reward = defaultdict(float)
dept_ae_reward = defaultdict(float)
qc_doctors = []  # [(dept, name, amount), ...]
ae_records = []  # [(dept, name, count, amount), ...]

for i, table in enumerate(doc.tables):
    if i == 15:  # Table 16 - 质控医生奖励
        for row in table.rows[1:]:
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) >= 3:
                name = cells[0]
                dept = cells[1]
                try:
                    amt = float(cells[2])
                    dept_qc_reward[dept] += amt
                    qc_doctors.append((dept, name, amt))
                except:
                    pass
    elif i == 16:  # Table 17 - 不良事件奖励
        for row in table.rows[1:]:
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) >= 4:
                name = cells[0]
                dept = cells[1]
                try:
                    cnt = int(cells[2])
                    amt = float(cells[3])
                    dept_ae_reward[dept] += amt
                    ae_records.append((dept, name, cnt, amt))
                except:
                    pass

# ===== 科室名称映射 =====
dept_mapping = {
    '儿科': '儿科', '耳鼻咽喉科': '耳鼻喉科', '耳鼻咽喉科盐港院区': '耳鼻喉科',
    '眼科': '眼科', '眼科盐港院区': '眼科',
    '肛肠科（中医科二）主院区': '中西医结合肛肠科', '肛肠科（中医科二）盐港院区': '中西医结合肛肠科',
    '中医科一': '中医科', '胸外科': '胸外科',
    '呼吸与危重症医学科': '呼吸内科', '普通外科': '普外科', '普通外科（盐港）': '普外科盐田',
    '神经内科': '神经内科', '神经外科': '神经外科',
    '针灸理疗科': '针灸推拿科', '中西医结合心血管内科': '中西医结合心血管内科',
    '肾内科': '肾内科', '消化内科': '消化内科', '泌尿外科': '泌尿外科',
    '泌尿外科（二）': '泌尿外科', '甲乳外科': '甲乳外科',
    '感染科-肝病门诊': '传染科', '感染性疾病科（含肝病）': '传染科',
    '内分泌科': '内分泌科', '产科': '产科',
    '血液内科、全科医学科': '血液内科', '肿瘤科': '肿瘤科',
    '骨科': '骨科', '骨科（盐港）': '骨科盐田',
    '康复医学科': '康复医学科', '重症医学科': '重症医学科',
    '新生儿科': '新生儿科', '老年医学科（综合内科）': '中西医结合老年病科',
    '皮肤科': '皮肤科', '皮肤科盐港院区': '皮肤科',
    '风湿免疫科': '风湿免疫科', '精神心理科': '精神心理科',
    '全科医学科': '全科医学科', '口腔科': '口腔科', '口腔科盐港院区': '口腔科',
    '营养科': '营养科', '急诊科盐田': '急诊科盐田', '急诊儿科': '儿科',
}

def aggregate_to_template(dept_dict):
    result = defaultdict(int)
    for src_dept, value in dept_dict.items():
        template_dept = dept_mapping.get(src_dept, src_dept)
        result[template_dept] += value
    return result

template_grade_b = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})
for src_dept, type_dict in dept_grade_b.items():
    template_dept = dept_mapping.get(src_dept, src_dept)
    for t in ['门诊', '环节', '终末']:
        template_grade_b[template_dept][t] += type_dict[t]

template_late_fine = aggregate_to_template(dept_late_fine)
template_excellent = aggregate_to_template(dept_excellent)
template_qc_reward = aggregate_to_template(dept_qc_reward)
template_ae_reward = aggregate_to_template(dept_ae_reward)

# ===== 生成新文件 =====
wb_new = openpyxl.Workbook()
ws1 = wb_new.active
ws1.title = 'Sheet1'
ws2 = wb_new.create_sheet('Sheet2')

# 样式设置
title_font = Font(bold=True, size=14)
header_font = Font(bold=True, size=11)
normal_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')

# ===== Sheet1: 完全参照原文件格式 =====

# 标题
ws1['A1'] = '2026年3月质管办质量奖罚汇总表'
ws1['A1'].font = title_font
ws1.merge_cells('A1:E1')
ws1.row_dimensions[1].height = 25

# (一) 病历奖惩
ws1['A2'] = '（一）病历奖惩'
ws1['A2'].font = header_font

# 表头
headers = ['科室', '门诊病历扣罚', '住院病历扣罚', '迟交病历扣罚', '优秀病历奖励']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# 病历奖惩数据 (26个科室)
dept_list_1 = [
    '儿科', '耳鼻喉科', '眼科', '中西医结合肛肠科', '中医科', '胸外科',
    '呼吸内科', '普外科', '神经内科', '中西医结合老年病科', '针灸推拿科',
    '中西医结合心血管内科', '肾内科', '消化内科', '泌尿外科', '普外科盐田',
    '妇科', '甲乳外科', '传染科', '内分泌科', '产科', '血液内科', '肿瘤科',
    '骨科盐田', '康复医学科', '重症医学科', '新生儿科', '神经外科', '骨科',
    '口腔科', '急诊科盐田', '皮肤科', '精神科'
]

row_idx = 4
data_start_row = row_idx
for dept in dept_list_1:
    ws1.cell(row=row_idx, column=1, value=dept).alignment = left_align
    gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    out_pen = gb['门诊'] * 100
    in_pen = gb['环节'] * 100 + gb['终末'] * 200
    late = int(template_late_fine.get(dept, 0))
    exc = template_excellent.get(dept, 0) * 300
    ws1.cell(row=row_idx, column=2, value=out_pen)
    ws1.cell(row=row_idx, column=3, value=in_pen)
    ws1.cell(row=row_idx, column=4, value=late)
    ws1.cell(row=row_idx, column=5, value=exc)
    for col in range(1, 6):
        ws1.cell(row=row_idx, column=col).border = thin_border
    row_idx += 1

data_end_row = row_idx - 1

# 汇总金额
ws1.cell(row=row_idx, column=1, value='汇总金额（元）').font = header_font
for col in range(2, 6):
    ws1.cell(row=row_idx, column=col).value = f'=SUM({chr(64+col)}{data_start_row}:{chr(64+col)}{data_end_row})'
    ws1.cell(row=row_idx, column=col).font = header_font
    ws1.cell(row=row_idx, column=col).border = thin_border
    ws1.cell(row=row_idx, column=col).alignment = center_align

row_idx += 1

# (二) 质控医生奖励
ws1.cell(row=row_idx, column=1, value='（二）质控医生奖励').font = header_font
row_idx += 1

# 表头
for col, h in enumerate(['科室', '姓名', '奖励金额'], 1):
    cell = ws1.cell(row=row_idx, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

row_idx += 1
qc_start_row = row_idx

# 质控医生数据 (按doc表中的人员)
qc_mapping = {
    '肿瘤科': ['刘湘园', '袁克华'], '眼科': ['李青蒨'], '急诊科': ['张伟艺', '黄建鑫'],
    '呼吸内科': ['陈祖华', '晏斌林'], '康复医学科': ['王敬芳', '涂文斌'], '血液内科': ['董航'],
    '针灸推拿科': ['王燕'], '中西医结合心血管内科': ['陈娟'], '皮肤科': ['李龙振'],
    '妇科': ['古晓珊', '王雪'], '内分泌科': ['左旋', '肖虎', '巫晓蓉'],
    '普外科': ['邹云东', '邱志东', '马达', '杨建桃', '刘兆艺'], '肾内科': ['赵飞', '潘志铣', '元华丹', '王佳林', '杜巧玲', '冯亚运'],
    '神经内科': ['邹社昌', '肖柏成'], '中西医结合肛肠科': ['陈飞云'],
    '甲乳外科': ['陈宝洁', '刘庆仪'], '骨科': ['李浩瑜', '刘鹤鸣', '周玉华', '翁雯梅'],
    '消化内科': ['张绍敏'], '中医科': ['曹丽军', '黄薇'], '产科': ['王敏', '李林娜', '柳调娟', '韩盼'],
    '传染科': ['蒋小民', '付小义', '刘恋'], '康复医学科盐田': ['廖伟东'],
    '骨科盐田': ['梁比记', '钟昌戎', '吴燕青'], '中西医结合老年病科': ['郑仲萍', '周赛赛', '黄小花'],
    '泌尿外科': ['李显文'], '耳鼻喉科': ['钟木生', '胡俊丽'], '新生儿科': ['张梦珍', '邱子莹'],
    '急诊科盐田': ['******', '李育英'], '神经外科': ['李燕霞'], '麻醉科': ['叶美', '张彩虹'],
    '检验科': ['吴意', '向惠英', '段虹如'], '口腔科': ['任丽婷', '南顺花'],
    '体检科': ['邓少芬'], '体检科盐田': ['刘冠俐'], '药剂科': ['许家眭'],
    '海涛社康': ['黄荫桂'], '沿港社康': ['刘静', '邓琼'], '明珠社康': ['李剑明'],
    '田东社康': ['邓玛丽'], '大梅沙社康': ['周婉'], '永安社康': ['张芳威'],
    '消毒供应中心': ['陈艳玲'],
}

for dept, names in qc_mapping.items():
    for name in names:
        ws1.cell(row=row_idx, column=1, value=dept).alignment = left_align
        ws1.cell(row=row_idx, column=2, value=name).alignment = left_align
        ws1.cell(row=row_idx, column=3, value=200)
        for col in range(1, 4):
            ws1.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

qc_end_row = row_idx - 1

# 汇总
ws1.cell(row=row_idx, column=1, value='汇总金额（元）').font = header_font
ws1.cell(row=row_idx, column=3).value = f'=SUM(C{qc_start_row}:C{qc_end_row})'
ws1.cell(row=row_idx, column=3).font = header_font
ws1.cell(row=row_idx, column=3).border = thin_border
ws1.cell(row=row_idx, column=3).alignment = center_align

row_idx += 1

# (三) 不良事件主动上报奖励
ws1.cell(row=row_idx, column=1, value='（三）不良事件主动上报奖励').font = header_font
row_idx += 1

for col, h in enumerate(['科室', '姓名', '上报例数', '奖励金额'], 1):
    cell = ws1.cell(row=row_idx, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

row_idx += 1
ae_start_row = row_idx

for dept, name, cnt, amt in ae_records:
    ws1.cell(row=row_idx, column=1, value=dept).alignment = left_align
    ws1.cell(row=row_idx, column=2, value=name).alignment = left_align
    ws1.cell(row=row_idx, column=3, value=cnt)
    ws1.cell(row=row_idx, column=4, value=amt)
    for col in range(1, 5):
        ws1.cell(row=row_idx, column=col).border = thin_border
    row_idx += 1

ae_end_row = row_idx - 1

# 汇总
ws1.cell(row=row_idx, column=1, value='汇总金额（元）').font = header_font
ws1.cell(row=row_idx, column=3).value = f'=SUM(C{ae_start_row}:C{ae_end_row})'
ws1.cell(row=row_idx, column=3).font = header_font
ws1.cell(row=row_idx, column=3).border = thin_border
ws1.cell(row=row_idx, column=3).alignment = center_align
ws1.cell(row=row_idx, column=4).value = f'=SUM(D{ae_start_row}:D{ae_end_row})'
ws1.cell(row=row_idx, column=4).font = header_font
ws1.cell(row=row_idx, column=4).border = thin_border
ws1.cell(row=row_idx, column=4).alignment = center_align

row_idx += 1

# (四) 单病种上报奖励
ws1.cell(row=row_idx, column=1, value='（四）单病种上报奖励').font = header_font
row_idx += 1

for col, h in enumerate(['科室', '上报例数（份）', '奖励金额（元）'], 1):
    cell = ws1.cell(row=row_idx, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

row_idx += 1
ws1.cell(row=row_idx, column=1, value='无').alignment = left_align
ws1.cell(row=row_idx, column=2, value=0)
ws1.cell(row=row_idx, column=3, value=0)
for col in range(1, 4):
    ws1.cell(row=row_idx, column=col).border = thin_border

row_idx += 1
ws1.cell(row=row_idx, column=1, value='汇总金额（元）').font = header_font
ws1.cell(row=row_idx, column=2).value = 0
ws1.cell(row=row_idx, column=3).value = 0
ws1.cell(row=row_idx, column=2).font = header_font
ws1.cell(row=row_idx, column=3).font = header_font

row_idx += 1

# 其他：
ws1.cell(row=row_idx, column=1, value='其他：').font = header_font
row_idx += 1

# 本月奖励合计 / 本月处罚合计 - 需要计算
# 计算各项汇总
total_reward = 0
total_penalty = 0

# 优秀病历奖励
for dept in dept_list_1:
    total_reward += template_excellent.get(dept, 0) * 300

# 质控医生奖励
total_qc = int(sum(v for v in template_qc_reward.values()))

# 不良事件奖励
total_ae = int(sum(v for v in template_ae_reward.values()))

# 单病种
total_disease = 0

# 计算病历扣罚
for dept in dept_list_1:
    gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    penalty = gb['门诊'] * 100 + gb['环节'] * 100 + gb['终末'] * 200 + int(template_late_fine.get(dept, 0))
    total_penalty += penalty

total_reward = total_qc + total_ae + sum(template_excellent.get(dept, 0) * 300 for dept in dept_list_1)

ws1.cell(row=row_idx, column=1, value='本月奖励合计').font = header_font
ws1.cell(row=row_idx, column=2, value=f'{total_reward}（元）').font = normal_font
ws1.cell(row=row_idx, column=3, value='本月处罚合计').font = header_font
ws1.cell(row=row_idx, column=4, value=f'{total_penalty}（元）').font = normal_font

row_idx += 1
ws1.cell(row=row_idx, column=1, value='制表人：何湛春').font = normal_font
ws1.cell(row=row_idx, column=3, value='审核人：王力刚').font = normal_font

row_idx += 1
ws1.cell(row=row_idx, column=1, value='领导意见：').font = header_font

# 调整列宽
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12

# ===== Sheet2 =====
ws2['A1'] = '2026年3月质管办质量奖罚汇总'
ws2['A1'].font = title_font
ws2.merge_cells('A1:F1')

# 表头
headers2 = ['科室', '病历扣罚金额', '优秀病历奖励金额', '质控医生奖励金额', '不良事件奖励', '单病种奖励']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# Sheet2的完整科室列表
dept_list_2 = [
    '传染科', '肾内科', '消化内科', '血液内科', '呼吸内科', '内分泌科',
    '肿瘤科', '重症医学科', '中西医结合心血管内科', '中西医结合肛肠科',
    '神经内科', '神经外科', '普外科', '普外科盐田', '骨科', '骨科盐田',
    '泌尿外科', '胸外科', '甲乳外科', '康复医学科', '针灸推拿科盐田',
    '眼科', '耳鼻喉科', '儿科', '新生儿科', '妇科', '产科', '中医科',
    '中西医结合老年病科', '急诊科', '急诊科（盐田）', '皮肤科', '营养科',
    '针灸推拿科', '口腔科', '麻醉科', '体检科', '检验科', '海涛社康',
    '明珠社康', '体检科盐田', '药剂科', '田东社康', '沿港社康', '大梅沙社康',
    '永安社康', '消毒供应中心', '其他'
]

row_idx = 3
for dept in dept_list_2:
    ws2.cell(row=row_idx, column=1, value=dept).alignment = left_align

    gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    penalty = gb['门诊'] * 100 + gb['环节'] * 100 + gb['终末'] * 200 + int(template_late_fine.get(dept, 0))
    exc = template_excellent.get(dept, 0) * 300
    qc = int(template_qc_reward.get(dept, 0))
    ae = int(template_ae_reward.get(dept, 0))

    ws2.cell(row=row_idx, column=2, value=penalty if penalty > 0 else None)
    ws2.cell(row=row_idx, column=3, value=exc if exc > 0 else None)

    # 质控医生奖励 - 显示人员名单
    if qc > 0:
        names = [n for d, n, a in qc_doctors if d == dept]
        ws2.cell(row=row_idx, column=4, value=f'{qc}（{"、".join(names)}）' if names else qc)
    else:
        ws2.cell(row=row_idx, column=4, value=None)

    ws2.cell(row=row_idx, column=5, value=ae if ae > 0 else None)
    ws2.cell(row=row_idx, column=6, value=0)

    for col in range(1, 7):
        ws2.cell(row=row_idx, column=col).border = thin_border

    row_idx += 1

# 汇总行
ws2.cell(row=row_idx, column=1, value='汇总金额（元）').font = header_font
ws2.cell(row=row_idx, column=2, value=f'{total_penalty}（扣）').font = normal_font
ws2.cell(row=row_idx, column=3, value=f'{sum(template_excellent.get(d, 0)*300 for d in dept_list_1)}（奖）').font = normal_font
ws2.cell(row=row_idx, column=4, value=f'{total_qc}（奖）').font = normal_font
ws2.cell(row=row_idx, column=5, value=f'{total_ae}（奖）').font = normal_font
ws2.cell(row=row_idx, column=6, value='0（奖）').font = normal_font

# 调整列宽
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12

# 保存
output_path = '2026年3月质管办奖罚通知（5月14日交经管）_final.xlsx'
wb_new.save(output_path)
print(f'已生成: {output_path}')

# 打印汇总
print(f'\n本月奖励合计: {total_reward}（元）')
print(f'本月处罚合计: {total_penalty}（元）')
print(f'质控医生奖励合计: {total_qc}（元）')
print(f'不良事件奖励合计: {total_ae}（元）')
print(f'优秀病历奖励合计: {sum(template_excellent.get(d, 0)*300 for d in dept_list_1)}（元）')