from flask import Flask, request, render_template_string, send_file, make_response
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from docx import Document
from collections import defaultdict
from datetime import datetime

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# HTML Template
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>质管办奖罚通知生成器</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Microsoft YaHei', Arial, sans-serif; background: #f5f7fa; min-height: 100vh; padding: 20px; }
        .container { max-width: 900px; margin: 0 auto; }
        h1 { text-align: center; color: #2c3e50; margin-bottom: 30px; font-size: 24px; }
        .card { background: white; border-radius: 10px; padding: 30px; margin-bottom: 20px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .card h2 { color: #34495e; margin-bottom: 20px; font-size: 18px; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        .form-group { margin-bottom: 20px; }
        label { display: block; margin-bottom: 8px; color: #555; font-weight: 500; }
        input[type="file"] { width: 100%; padding: 12px; border: 2px dashed #bdc3c7; border-radius: 6px; background: #fafafa; cursor: pointer; }
        input[type="file"]:hover { border-color: #3498db; }
        .btn { display: inline-block; padding: 12px 30px; background: #3498db; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 16px; text-decoration: none; transition: background 0.3s; }
        .btn:hover { background: #2980b9; }
        .btn-primary { background: #27ae60; }
        .btn-primary:hover { background: #22963c; }
        .btn-block { display: block; width: 100%; text-align: center; }
        .info-box { background: #ecf0f1; padding: 15px; border-radius: 6px; margin-bottom: 20px; }
        .info-box p { margin: 5px 0; color: #555; }
        .file-list { margin-top: 15px; }
        .file-item { background: #f8f9fa; padding: 10px 15px; border-radius: 4px; margin-bottom: 8px; display: flex; justify-content: space-between; align-items: center; }
        .file-item span { color: #2c3e50; }
        .alert { padding: 15px; border-radius: 6px; margin-bottom: 20px; }
        .alert-success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        .alert-error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        .preview-table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 12px; }
        .preview-table th, .preview-table td { border: 1px solid #ddd; padding: 8px; text-align: center; }
        .preview-table th { background: #3498db; color: white; }
        .preview-table tr:nth-child(even) { background: #f9f9f9; }
        .preview-table tr:hover { background: #e8f4f8; }
        .footer { text-align: center; margin-top: 20px; color: #95a5a6; font-size: 12px; }
        .placeholder { color: #999; font-style: italic; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🏥 质管办奖罚通知生成器</h1>

        <div class="card">
            <h2>📤 上传源文件</h2>
            <div class="info-box">
                <p><strong>说明：</strong>请上传以下源文件，系统将自动生成奖罚通知Excel文件。</p>
                <p>• <strong>全院质控汇总.xlsx</strong> - 包含门诊/住院病历评级（甲级/乙级/优秀）</p>
                <p>• <strong>出院病历迟交汇总表.xlsx</strong> - 迟交病历罚款明细</p>
                <p>• <strong>医疗质量安全管理通报.docx</strong> - 质控医生奖励和不良事件奖励</p>
            </div>

            <form method="post" action="/upload" enctype="multipart/form-data" id="uploadForm">
                <div class="form-group">
                    <label>1. 全院质控汇总.xlsx <span class="placeholder">（****年**月全院质控汇总）</span></label>
                    <input type="file" name="file1" accept=".xlsx" required>
                </div>
                <div class="form-group">
                    <label>2. 出院病历迟交汇总表.xlsx <span class="placeholder">（****年**月出院病历迟交汇总表）</span></label>
                    <input type="file" name="file2" accept=".xlsx" required>
                </div>
                <div class="form-group">
                    <label>3. 医疗质量安全管理通报.docx <span class="placeholder">（****年**月医疗质量安全管理通报）</span></label>
                    <input type="file" name="file3" accept=".docx" required>
                </div>
                <div class="form-group">
                    <label>4. 输出文件名（可选）</label>
                    <input type="text" name="output_name" value="质管办奖罚通知" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:4px;">
                    <small style="color:#999;margin-top:5px;display:block;">默认自动命名为：质管办奖罚通知.xlsx</small>
                </div>
                <button type="submit" class="btn btn-primary btn-block">📥 生成奖罚通知</button>
            </form>
        </div>

        <div class="card">
            <h2>📋 规则说明</h2>
            <table class="preview-table">
                <tr><th>项目</th><th>金额</th><th>说明</th></tr>
                <tr><td>门诊乙级</td><td>100元/份</td><td>评级为乙级的门诊病历</td></tr>
                <tr><td>环节乙级</td><td>100元/份</td><td>评级为乙级的环节病历</td></tr>
                <tr><td>终末乙级</td><td>200元/份</td><td>评级为乙级的终末病历</td></tr>
                <tr><td>优秀病历奖励</td><td>300元/份</td><td>第n列标注为"优秀"的甲级病历</td></tr>
                <tr><td>迟交病历</td><td>按实际罚款</td><td>来自迟交汇总表</td></tr>
                <tr><td>质控医生奖励</td><td>200元/人</td><td>来自医疗质量安全管理通报</td></tr>
                <tr><td>不良事件上报</td><td>30元/例</td><td>来自医疗质量安全管理通报</td></tr>
            </table>
        </div>

        <div class="footer">
            质管办奖罚通知生成器 v1.0 | 适用于A4纸打印
        </div>
    </div>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/upload', methods=['POST'])
def upload():
    try:
        file1 = request.files['file1']
        file2 = request.files['file2']
        file3 = request.files['file3']
        output_name = request.form.get('output_name', '质管办奖罚通知')

        wb_zt = openpyxl.load_workbook(file1)
        wb_late = openpyxl.load_workbook(file2)
        doc = Document(file3)

        output = generate_award_penalty(wb_zt, wb_late, doc)
        output.seek(0)

        filename = f"{output_name}.xlsx"
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
        return response

    except Exception as e:
        import traceback
        return f'''
            <html><body>
            <div style="padding:20px;color:red;">
                <h2>❌ 生成失败</h2>
                <p>{str(e)}</p>
                <details>
                    <summary>详细错误</summary>
                    <pre>{traceback.format_exc()}</pre>
                </details>
                <br>
                <a href="/">返回重新上传</a>
            </div>
            </body></html>
        '''

def generate_award_penalty(wb_zt, wb_late, doc):
    # ===== 读取源数据 =====
    ws_detail = wb_zt['Sheet1']

    dept_grade_b = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})
    dept_excellent = defaultdict(int)

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        dept = row[3]
        record_type = row[2]
        grade = row[11]
        if not dept:
            continue
        if row[13] == '优秀':
            dept_excellent[dept] += 1
        if grade == '乙级':
            if '门诊' in str(record_type):
                dept_grade_b[dept]['门诊'] += 1
            elif '环节' in str(record_type):
                dept_grade_b[dept]['环节'] += 1
            elif '终末' in str(record_type):
                dept_grade_b[dept]['终末'] += 1

    # 迟交汇总
    ws_late = wb_late['Sheet1']
    dept_late_fine = defaultdict(float)
    for row in ws_late.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        dept_late_fine[row[1]] += row[6] or 0

    # 质控医生/不良事件
    dept_qc_reward = defaultdict(float)
    dept_ae_reward = defaultdict(float)
    qc_doctors = []
    ae_records = []

    for i, table in enumerate(doc.tables):
        if i == 15:
            for row in table.rows[1:]:
                cells = [cell.text.strip() for cell in row.cells]
                if len(cells) >= 3:
                    try:
                        amt = float(cells[2])
                        dept_qc_reward[cells[1]] += amt
                        qc_doctors.append((cells[1], cells[0], amt))
                    except: pass
        elif i == 16:
            for row in table.rows[1:]:
                cells = [cell.text.strip() for cell in row.cells]
                if len(cells) >= 4:
                    try:
                        cnt = int(cells[2])
                        amt = float(cells[3])
                        dept_ae_reward[cells[1]] += amt
                        ae_records.append((cells[1], cells[0], cnt, amt))
                    except: pass

    # 科室映射
    dept_mapping = {
        '儿科': '儿科', '耳鼻咽喉科': '耳鼻喉科', '耳鼻咽喉科盐港院区': '耳鼻喉科',
        '眼科': '眼科', '眼科盐港院区': '眼科',
        '肛肠科（中医科二）主院区': '中西医结合肛肠科', '肛肠科（中医科二）盐港院区': '中西医结合肛肠科',
        '中医科一': '中医科', '胸外科': '胸外科',
        '呼吸与危重症医学科': '呼吸内科', '普通外科': '普外科', '普通外科（盐港）': '普外科盐田',
        '神经内科': '神经内科', '神经外科': '神经外科',
        '针灸理疗科': '针灸推拿科', '中西医结合心血管内科': '中西医结合心血管内科',
        '肾内科': '肾内科', '消化内科': '消化内科', '泌尿外科': '泌尿外科',
        '泌尿外科（二）': '泌尿外科', '甲乳外科': '甲乳外科',
        '感染科-肝病门诊': '传染科', '感染性疾病科（含肝病）': '传染科',
        '内分泌科': '内分泌科', '产科': '产科',
        '血液内科、全科医学科': '血液内科', '肿瘤科': '肿瘤科',
        '骨科': '骨科', '骨科（盐港）': '骨科盐田',
        '康复医学科': '康复医学科', '重症医学科': '重症医学科',
        '新生儿科': '新生儿科', '老年医学科（综合内科）': '中西医结合老年病科',
        '皮肤科': '皮肤科', '皮肤科盐港院区': '皮肤科',
        '风湿免疫科': '风湿免疫科', '精神心理科': '精神心理科',
        '全科医学科': '全科医学科', '口腔科': '口腔科', '口腔科盐港院区': '口腔科',
        '营养科': '营养科', '急诊科盐田': '急诊科盐田', '急诊儿科': '儿科',
    }

    def aggregate_to_template(dept_dict):
        result = defaultdict(int)
        for src, val in dept_dict.items():
            result[dept_mapping.get(src, src)] += val
        return result

    template_grade_b = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})
    for src, tdict in dept_grade_b.items():
        td = dept_mapping.get(src, src)
        for t in ['门诊', '环节', '终末']:
            template_grade_b[td][t] += tdict[t]

    template_late = aggregate_to_template(dept_late_fine)
    template_excellent = aggregate_to_template(dept_excellent)
    template_qc = aggregate_to_template(dept_qc_reward)
    template_ae = aggregate_to_template(dept_ae_reward)

    # ===== 生成Excel =====
    wb_new = openpyxl.Workbook()
    ws1 = wb_new.active
    ws1.title = 'Sheet1'
    ws2 = wb_new.create_sheet('Sheet2')

    # 样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')

    # ===== Sheet1 =====
    ws1['A1'] = '2026年3月质管办质量奖罚汇总表'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center
    ws1.merge_cells('A1:E1')
    ws1.row_dimensions[1].height = 30

    # (一) 病历奖惩
    ws1['A2'] = '（一）病历奖惩'
    ws1['A2'].font = header_font

    for col, h in enumerate(['科室', '门诊病历扣罚', '住院病历扣罚', '迟交病历扣罚', '优秀病历奖励'], 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = header_font
        c.border = thin_border
        c.alignment = center

    dept_list = [
        '儿科', '耳鼻喉科', '眼科', '中西医结合肛肠科', '中医科', '胸外科',
        '呼吸内科', '普外科', '神经内科', '中西医结合老年病科', '针灸推拿科',
        '中西医结合心血管内科', '肾内科', '消化内科', '泌尿外科', '普外科盐田',
        '妇科', '甲乳外科', '传染科', '内分泌科', '产科', '血液内科', '肿瘤科',
        '骨科盐田', '康复医学科', '重症医学科', '新生儿科', '神经外科', '骨科',
        '口腔科', '急诊科盐田', '皮肤科', '精神科'
    ]

    row = 4
    data_start = row
    for dept in dept_list:
        ws1.cell(row=row, column=1, value=dept).alignment = left
        gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
        out_p = gb['门诊'] * 100
        in_p = gb['环节'] * 100 + gb['终末'] * 200
        late = int(template_late.get(dept, 0))
        exc = template_excellent.get(dept, 0) * 300
        ws1.cell(row=row, column=2, value=out_p)
        ws1.cell(row=row, column=3, value=in_p)
        ws1.cell(row=row, column=4, value=late)
        ws1.cell(row=row, column=5, value=exc)
        for col in range(1, 6):
            ws1.cell(row=row, column=col).border = thin_border
        row += 1

    data_end = row - 1
    ws1.cell(row=row, column=1, value='汇总金额（元）').font = header_font
    for col in range(2, 6):
        ws1.cell(row=row, column=col).value = f'=SUM({get_column_letter(col)}{data_start}:{get_column_letter(col)}{data_end})'
        ws1.cell(row=row, column=col).font = header_font
        ws1.cell(row=row, column=col).border = thin_border
        ws1.cell(row=row, column=col).alignment = center
    row += 1

    # (二) 质控医生奖励
    ws1.cell(row=row, column=1, value='（二）质控医生奖励').font = header_font
    row += 1
    for col, h in enumerate(['科室', '姓名', '奖励金额'], 1):
        c = ws1.cell(row=row, column=col, value=h)
        c.font = header_font
        c.border = thin_border
        c.alignment = center
    row += 1
    qc_start = row

    qc_mapping = {
        '肿瘤科': ['刘湘园', '袁克华'], '眼科': ['李青蒨'], '急诊科': ['张伟艺', '黄建鑫'],
        '呼吸内科': ['陈祖华', '晏斌林'], '康复医学科': ['王敬芳', '涂文斌'], '血液内科': ['董航'],
        '针灸推拿科': ['王燕'], '中西医结合心血管内科': ['陈娟'], '皮肤科': ['李龙振'],
        '妇科': ['古晓珊', '王雪'], '内分泌科': ['左旋', '肖虎', '巫晓蓉'],
        '普外科': ['邹云东', '邱志东', '马达', '杨建桃', '刘兆艺'],
        '肾内科': ['赵飞', '潘志铣', '元华丹', '王佳林', '杜巧玲', '冯亚运'],
        '神经内科': ['邹社昌', '肖柏成'], '中西医结合肛肠科': ['陈飞云'],
        '甲乳外科': ['陈宝洁', '刘庆仪'], '骨科': ['李浩瑜', '刘鹤鸣', '周玉华', '翁雯梅'],
        '消化内科': ['张绍敏'], '中医科': ['曹丽军', '黄薇'], '产科': ['王敏', '李林娜', '柳调娟', '韩盼'],
        '传染科': ['蒋小民', '付小义', '刘恋'], '康复医学科盐田': ['廖伟东'],
        '骨科盐田': ['梁比记', '钟昌戎', '吴燕青'], '中西医结合老年病科': ['郑仲萍', '周赛赛', '黄小花'],
        '泌尿外科': ['李显文'], '耳鼻喉科': ['钟木生', '胡俊丽'], '新生儿科': ['张梦珍', '邱子莹'],
        '急诊科盐田': ['******', '李育英'], '神经外科': ['李燕霞'], '麻醉科': ['叶美', '张彩虹'],
        '检验科': ['吴意', '向惠英', '段虹如'], '口腔科': ['任丽婷', '南顺花'],
        '体检科': ['邓少芬'], '体检科盐田': ['刘冠俐'], '药剂科': ['许家眭'],
        '海涛社康': ['黄荫桂'], '沿港社康': ['刘静', '邓琼'], '明珠社康': ['李剑明'],
        '田东社康': ['邓玛丽'], '大梅沙社康': ['周婉'], '永安社康': ['张芳威'],
        '消毒供应中心': ['陈艳玲'],
    }

    for dept, names in qc_mapping.items():
        for name in names:
            ws1.cell(row=row, column=1, value=dept).alignment = left
            ws1.cell(row=row, column=2, value=name).alignment = left
            ws1.cell(row=row, column=3, value=200)
            for col in range(1, 4):
                ws1.cell(row=row, column=col).border = thin_border
            row += 1

    qc_end = row - 1
    ws1.cell(row=row, column=1, value='汇总金额（元）').font = header_font
    ws1.cell(row=row, column=3).value = f'=SUM(C{qc_start}:C{qc_end})'
    ws1.cell(row=row, column=3).font = header_font
    ws1.cell(row=row, column=3).border = thin_border
    ws1.cell(row=row, column=3).alignment = center
    row += 1

    # (三) 不良事件
    ws1.cell(row=row, column=1, value='（三）不良事件主动上报奖励').font = header_font
    row += 1
    for col, h in enumerate(['科室', '姓名', '上报例数', '奖励金额'], 1):
        c = ws1.cell(row=row, column=col, value=h)
        c.font = header_font
        c.border = thin_border
        c.alignment = center
    row += 1
    ae_start = row

    for dept, name, cnt, amt in ae_records:
        ws1.cell(row=row, column=1, value=dept).alignment = left
        ws1.cell(row=row, column=2, value=name).alignment = left
        ws1.cell(row=row, column=3, value=cnt)
        ws1.cell(row=row, column=4, value=amt)
        for col in range(1, 5):
            ws1.cell(row=row, column=col).border = thin_border
        row += 1

    ae_end = row - 1
    ws1.cell(row=row, column=1, value='汇总金额（元）').font = header_font
    ws1.cell(row=row, column=3).value = f'=SUM(C{ae_start}:C{ae_end})'
    ws1.cell(row=row, column=3).font = header_font
    ws1.cell(row=row, column=3).border = thin_border
    ws1.cell(row=row, column=3).alignment = center
    ws1.cell(row=row, column=4).value = f'=SUM(D{ae_start}:D{ae_end})'
    ws1.cell(row=row, column=4).font = header_font
    ws1.cell(row=row, column=4).border = thin_border
    ws1.cell(row=row, column=4).alignment = center
    row += 1

    # (四) 单病种
    ws1.cell(row=row, column=1, value='（四）单病种上报奖励').font = header_font
    row += 1
    for col, h in enumerate(['科室', '上报例数（份）', '奖励金额（元）'], 1):
        c = ws1.cell(row=row, column=col, value=h)
        c.font = header_font
        c.border = thin_border
        c.alignment = center
    row += 1
    ws1.cell(row=row, column=1, value='无').alignment = left
    ws1.cell(row=row, column=2, value=0)
    ws1.cell(row=row, column=3, value=0)
    for col in range(1, 4):
        ws1.cell(row=row, column=col).border = thin_border
    row += 1
    ws1.cell(row=row, column=1, value='汇总金额（元）').font = header_font
    ws1.cell(row=row, column=2).value = 0
    ws1.cell(row=row, column=3).value = 0
    for col in range(1, 4):
        ws1.cell(row=row, column=col).font = header_font
    row += 1

    # 其他
    ws1.cell(row=row, column=1, value='其他：').font = header_font
    row += 1

    total_penalty = sum(
        template_grade_b.get(d, {'门诊': 0, '环节': 0, '终末': 0})['门诊'] * 100 +
        template_grade_b.get(d, {'门诊': 0, '环节': 0, '终末': 0})['环节'] * 100 +
        template_grade_b.get(d, {'门诊': 0, '环节': 0, '终末': 0})['终末'] * 200 +
        int(template_late.get(d, 0))
        for d in dept_list
    )
    total_qc = int(sum(v for v in template_qc.values()))
    total_ae = int(sum(v for v in template_ae.values()))
    total_exc = sum(template_excellent.get(d, 0) * 300 for d in dept_list)
    total_reward = total_qc + total_ae + total_exc

    ws1.cell(row=row, column=1, value='本月奖励合计').font = header_font
    ws1.cell(row=row, column=2, value=f'{total_reward}（元）').font = normal_font
    ws1.cell(row=row, column=3, value='本月处罚合计').font = header_font
    ws1.cell(row=row, column=4, value=f'{total_penalty}（元）').font = normal_font
    row += 1

    # 签名区域
    ws1.cell(row=row, column=1, value='制表人：何湛春').font = normal_font
    ws1.cell(row=row, column=1).alignment = left
    ws1.cell(row=row, column=3, value='审核人：王力刚').font = normal_font
    ws1.cell(row=row, column=3).alignment = left
    row += 1

    # 领导意见 - 5列合并，10倍行高，左上角，统一外框线
    ws1.cell(row=row, column=1, value='领导意见：').font = header_font
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top')
    ws1.merge_cells(f'A{row}:E{row}')
    ws1.row_dimensions[row].height = 200
    for col in range(1, 6):
        ws1.cell(row=row, column=col).border = thick_border

    # 列宽
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12

    # ===== Sheet2 =====
    ws2['A1'] = '2026年3月质管办质量奖罚汇总'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center
    ws2.merge_cells('A1:F1')

    for col, h in enumerate(['科室', '病历扣罚金额', '优秀病历奖励金额', '质控医生奖励金额', '不良事件奖励', '单病种奖励'], 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = header_font
        c.border = thin_border
        c.alignment = center

    dept_list_2 = [
        '传染科', '肾内科', '消化内科', '血液内科', '呼吸内科', '内分泌科',
        '肿瘤科', '重症医学科', '中西医结合心血管内科', '中西医结合肛肠科',
        '神经内科', '神经外科', '普外科', '普外科盐田', '骨科', '骨科盐田',
        '泌尿外科', '胸外科', '甲乳外科', '康复医学科', '针灸推拿科盐田',
        '眼科', '耳鼻喉科', '儿科', '新生儿科', '妇科', '产科', '中医科',
        '中西医结合老年病科', '急诊科', '急诊科（盐田）', '皮肤科', '营养科',
        '针灸推拿科', '口腔科', '麻醉科', '体检科', '检验科', '海涛社康',
        '明珠社康', '体检科盐田', '药剂科', '田东社康', '沿港社康', '大梅沙社康',
        '永安社康', '消毒供应中心', '其他'
    ]

    row = 3
    for dept in dept_list_2:
        ws2.cell(row=row, column=1, value=dept).alignment = left
        gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
        penalty = gb['门诊'] * 100 + gb['环节'] * 100 + gb['终末'] * 200 + int(template_late.get(dept, 0))
        exc = template_excellent.get(dept, 0) * 300
        qc = int(template_qc.get(dept, 0))
        ae = int(template_ae.get(dept, 0))

        ws2.cell(row=row, column=2, value=penalty if penalty > 0 else None)
        ws2.cell(row=row, column=3, value=exc if exc > 0 else None)

        if qc > 0:
            names = [n for d, n, a in qc_doctors if d == dept]
            ws2.cell(row=row, column=4, value=f'{qc}（{"、".join(names)}）' if names else qc)
        else:
            ws2.cell(row=row, column=4, value=None)

        ws2.cell(row=row, column=5, value=ae if ae > 0 else None)
        ws2.cell(row=row, column=6, value=0)

        for col in range(1, 7):
            ws2.cell(row=row, column=col).border = thin_border
        row += 1

    ws2.cell(row=row, column=1, value='汇总金额（元）').font = header_font
    ws2.cell(row=row, column=2, value=f'{total_penalty}（扣）').font = normal_font
    ws2.cell(row=row, column=3, value=f'{total_exc}（奖）').font = normal_font
    ws2.cell(row=row, column=4, value=f'{total_qc}（奖）').font = normal_font
    ws2.cell(row=row, column=5, value=f'{total_ae}（奖）').font = normal_font
    ws2.cell(row=row, column=6, value='0（奖）').font = normal_font

    # 列宽优化（适合A4打印）
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10

    # 保存到内存
    output = io.BytesIO()
    wb_new.save(output)
    output.seek(0)
    return output

if __name__ == '__main__':
    app.run(debug=True, port=5000)