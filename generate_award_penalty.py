import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from docx import Document
from collections import defaultdict

# ===== 读取源数据 =====

# 1. 读取3月全院质控汇总 - Sheet1 (详细数据)
wb_src1 = openpyxl.load_workbook('3月全院质控汇总.xlsx')
ws_detail = wb_src1['Sheet1']

# 统计各科室病历问题数
dept_grade_a = defaultdict(int)             # 甲级数
dept_grade_b = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})  # 乙级数(按病历类型)
dept_excellent = defaultdict(int)           # 优秀病历数（第n列标注为"优秀"）

for row in ws_detail.iter_rows(min_row=2, values_only=True):
    dept = row[3]   # 检查科室
    record_type = row[2]  # 病历类型
    issues = row[10]  # 存在问题
    grade = row[11]  # 病历评级

    if not dept:
        continue

    # 统计甲级
    if grade == '甲级':
        dept_grade_a[dept] += 1

    # 统计优秀病历（第13列标注为"优秀"）
    if row[13] == '优秀':
        dept_excellent[dept] += 1

    # 统计乙级(按病历类型)
    if grade == '乙级':
        if '门诊' in str(record_type):
            dept_grade_b[dept]['门诊'] += 1
        elif '环节' in str(record_type):
            dept_grade_b[dept]['环节'] += 1
        elif '终末' in str(record_type):
            dept_grade_b[dept]['终末'] += 1

# 2. 读取3月全院质控汇总 - Sheet2 (汇总数据)
ws_summary = wb_src1['Sheet2']
dept_total = {}  # 科室总检查数
dept_jiayi = {}   # 甲级数
dept_yiji = {}    # 乙级数

for row in ws_summary.iter_rows(min_row=5, values_only=True):
    if not row[0] or row[0] == '总计':
        continue
    dept = row[0]
    dept_jiayi[dept] = row[1] or 0
    dept_yiji[dept] = row[2] or 0
    dept_total[dept] = row[3] or 0

# 3. 读取迟交汇总
wb_late = openpyxl.load_workbook('2026年3月出院病历迟交汇总表.xlsx')
ws_late = wb_late['Sheet1']

dept_late_fine = defaultdict(float)  # 迟交罚款

for row in ws_late.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    dept = row[1]
    fine = row[6] or 0
    dept_late_fine[dept] += fine

# 4. 读取质控医生奖励和不良事件奖励
doc = Document('2026年03月医疗质量安全管理通报.docx')

dept_qc_reward = defaultdict(float)  # 质控医生奖励(按科室)
dept_ae_reward = defaultdict(float)   # 不良事件奖励(按科室)

# Table 16 (index 15): 质控医生奖励 - ['姓名', '科室', '奖励金额']
# Table 17 (index 16): 不良事件奖励 - ['姓名', '科室', '上报例数', '奖励金额']

for i, table in enumerate(doc.tables):
    if i == 15:  # Table 16 - 质控医生奖励
        for row in table.rows[1:]:  # 跳过表头
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) >= 3:
                name = cells[0]
                dept = cells[1]
                amount = cells[2]
                try:
                    dept_qc_reward[dept] += float(amount)
                except:
                    pass
    elif i == 16:  # Table 17 - 不良事件奖励
        for row in table.rows[1:]:  # 跳过表头
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) >= 4:
                name = cells[0]
                dept = cells[1]
                amount = cells[3]  # 第4列是奖励金额
                try:
                    dept_ae_reward[dept] += float(amount)
                except:
                    pass

print('=== 质控医生奖励(科室) ===')
for dept, amt in sorted(dept_qc_reward.items()):
    print(f'{dept}: {amt}')

print('\n=== 不良事件奖励(科室) ===')
for dept, amt in sorted(dept_ae_reward.items()):
    print(f'{dept}: {amt}')

# ===== 科室名称映射 (源数据 -> 模板) =====
dept_mapping = {
    '儿科': '儿科',
    '耳鼻咽喉科': '耳鼻喉科',
    '耳鼻咽喉科盐港院区': '耳鼻喉科',  # 合并
    '眼科': '眼科',
    '眼科盐港院区': '眼科',  # 合并
    '肛肠科（中医科二）主院区': '中西医结合肛肠科',
    '肛肠科（中医科二）盐港院区': '中西医结合肛肠科',
    '中医科一': '中医科',
    '胸外科': '胸外科',
    '呼吸与危重症医学科': '呼吸内科',
    '普通外科': '普外科',
    '普通外科（盐港）': '普外科盐田',
    '神经内科': '神经内科',
    '神经外科': '神经外科',
    '针灸理疗科': '针灸推拿科',
    '中西医结合心血管内科': '中西医结合心血管内科',
    '肾内科': '肾内科',
    '消化内科': '消化内科',
    '泌尿外科': '泌尿外科',
    '泌尿外科（二）': '泌尿外科',
    '甲乳外科': '甲乳外科',
    '感染科-肝病门诊': '传染科',
    '感染性疾病科（含肝病）': '传染科',
    '内分泌科': '内分泌科',
    '产科': '产科',
    '血液内科、全科医学科': '血液内科',
    '肿瘤科': '肿瘤科',
    '骨科': '骨科',
    '骨科（盐港）': '骨科盐田',
    '康复医学科': '康复医学科',
    '重症医学科': '重症医学科',
    '新生儿科': '新生儿科',
    '老年医学科（综合内科）': '中西医结合老年病科',
    '皮肤科': '皮肤科',
    '皮肤科盐港院区': '皮肤科',
    '风湿免疫科': '风湿免疫科',
    '精神心理科': '精神心理科',
    '全科医学科': '全科医学科',
    '口腔科': '口腔科',
    '口腔科盐港院区': '口腔科',
    '营养科': '营养科',
    '急诊科盐田': '急诊科',
    '急诊儿科': '急诊儿科',
}

# ===== 计算各模板科室的汇总数据 =====
template_depts = [
    '儿科', '耳鼻喉科', '眼科', '中西医结合肛肠科', '中医科',
    '胸外科', '呼吸内科', '普外科', '神经内科', '中西医结合老年病科',
    '针灸推拿科', '中西医结合心血管内科', '肾内科', '消化内科', '泌尿外科',
    '普外科盐田', '妇科', '甲乳外科', '传染科', '内分泌科',
    '产科', '血液内科', '肿瘤科', '骨科盐田', '康复医学科', '重症医学科', '新生儿科'
]

# 聚合到模板科室
def aggregate_to_template(dept_dict):
    result = defaultdict(int)
    for src_dept, value in dept_dict.items():
        template_dept = dept_mapping.get(src_dept, src_dept)
        result[template_dept] += value
    return result

template_late_fine = aggregate_to_template(dept_late_fine)
template_excellent = aggregate_to_template(dept_excellent)

# 聚合乙级(按类型)
def aggregate_grade_b():
    result = defaultdict(lambda: {'门诊': 0, '环节': 0, '终末': 0})
    for src_dept, type_dict in dept_grade_b.items():
        template_dept = dept_mapping.get(src_dept, src_dept)
        for t in ['门诊', '环节', '终末']:
            result[template_dept][t] += type_dict[t]
    return result

template_grade_b = aggregate_grade_b()

# 聚合质控医生奖励和不良事件奖励
template_qc_reward = aggregate_to_template(dept_qc_reward)
template_ae_reward = aggregate_to_template(dept_ae_reward)

# ===== 生成新文件 =====
wb_new = openpyxl.Workbook()
ws1 = wb_new.active
ws1.title = 'Sheet1'
ws2 = wb_new.create_sheet('Sheet2')

# 设置样式
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ===== Sheet1: 病历奖惩 =====
ws1['A1'] = '2026年3月质管办质量奖罚汇总表'
ws1['A1'].font = title_font
ws1.merge_cells('A1:E1')

ws1['A2'] = '（一）病历奖惩'
ws1['A2'].font = header_font

# 表头
headers1 = ['科室', '门诊病历扣罚', '住院病历扣罚', '迟交病历扣罚', '优秀病历奖励']
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# 计算扣罚金额
# 问题数 * 50元/个 + 乙级罚款(门诊/环节100元/份, 终末200元/份)
def calc_penalty(issues):
    return 0  # 问题扣罚已取消

def calc_grade_b_penalty(grade_b_dict):
    return grade_b_dict['门诊'] * 100 + grade_b_dict['环节'] * 100 + grade_b_dict['终末'] * 200

# 数据行
for row_idx, dept in enumerate(template_depts, 4):
    ws1.cell(row=row_idx, column=1, value=dept).alignment = left_align

    # 门诊病历扣罚 = 门诊乙级罚款 (100元/份)
    # 住院病历扣罚 = 环节乙级罚款(100元/份) + 终末乙级罚款(200元/份)
    grade_b = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    total_outpatient_penalty = grade_b['门诊'] * 100
    total_inpatient_penalty = grade_b['环节'] * 100 + grade_b['终末'] * 200

    late_fine = int(template_late_fine.get(dept, 0))
    excellent_reward = template_excellent.get(dept, 0) * 300

    ws1.cell(row=row_idx, column=2, value=total_outpatient_penalty)
    ws1.cell(row=row_idx, column=3, value=total_inpatient_penalty)
    ws1.cell(row=row_idx, column=4, value=late_fine)
    ws1.cell(row=row_idx, column=5, value=excellent_reward)

    for col in range(1, 6):
        ws1.cell(row=row_idx, column=col).border = thin_border

# ===== Sheet2: 详细奖罚 =====
ws2['A1'] = '2026年3月质管办质量奖罚汇总'
ws2['A1'].font = title_font
ws2.merge_cells('A1:F1')

# 表头
headers2 = ['科室', '病历扣罚金额', '优秀病历奖励金额', '质控医生奖励金额', '不良事件奖励', '单病种奖励']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# Sheet2: 病历扣罚 = 门诊乙级罚款 + 环节乙级罚款 + 终末乙级罚款 + 迟交扣罚
for row_idx, dept in enumerate(template_depts, 3):
    ws2.cell(row=row_idx, column=1, value=dept).alignment = left_align

    grade_b = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    grade_b_penalty = grade_b['门诊'] * 100 + grade_b['环节'] * 100 + grade_b['终末'] * 200
    late_fine = int(template_late_fine.get(dept, 0))
    excellent_reward = template_excellent.get(dept, 0) * 300
    qc_reward = template_qc_reward.get(dept, 0)
    ae_reward = template_ae_reward.get(dept, 0)

    total_penalty = grade_b_penalty + late_fine

    ws2.cell(row=row_idx, column=2, value=total_penalty if total_penalty > 0 else None)
    ws2.cell(row=row_idx, column=3, value=excellent_reward if excellent_reward > 0 else None)
    ws2.cell(row=row_idx, column=4, value=int(qc_reward) if qc_reward > 0 else None)
    ws2.cell(row=row_idx, column=5, value=int(ae_reward) if ae_reward > 0 else None)
    ws2.cell(row=row_idx, column=6, value=0)

    for col in range(1, 7):
        ws2.cell(row=row_idx, column=col).border = thin_border

# 调整列宽
ws1.column_dimensions['A'].width = 20
ws2.column_dimensions['A'].width = 20

# 保存
output_path = '2026年3月质管办奖罚通知（5月14日交经管）_v3.xlsx'
wb_new.save(output_path)
print(f'已生成: {output_path}')

# 打印汇总供核对
print('\n=== Sheet1 数据汇总 ===')
print(f'{"科室":<15} {"门诊扣罚":>10} {"住院扣罚":>10} {"迟交扣罚":>10} {"优秀奖励":>10}')
for dept in template_depts:
    grade_b = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    out_p = grade_b['门诊'] * 100
    in_p = grade_b['环节'] * 100 + grade_b['终末'] * 200
    late = int(template_late_fine.get(dept, 0))
    exc = template_excellent.get(dept, 0) * 300
    if out_p > 0 or in_p > 0 or late > 0 or exc > 0:
        print(f'{dept:<15} {out_p:>10} {in_p:>10} {late:>10} {exc:>10}')

print('\n=== 乙级病历明细 ===')
for dept in template_depts:
    gb = template_grade_b.get(dept, {'门诊': 0, '环节': 0, '终末': 0})
    if gb['门诊'] > 0 or gb['环节'] > 0 or gb['终末'] > 0:
        print(f'{dept}: 门诊乙级={gb["门诊"]}, 环节乙级={gb["环节"]}, 终末乙级={gb["终末"]}')

print('\n=== 优秀病历明细 ===')
for dept in template_depts:
    exc = template_excellent.get(dept, 0)
    if exc > 0:
        print(f'{dept}: {exc}份')
